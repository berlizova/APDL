import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Define the path to the solve.out file
solve_out_path = r'C:\Users\tebe\Documents\test\solve.out'  # Update the path to solve.out

# Define the output Excel file path (same folder as solve.out)
output_excel_path = os.path.join(os.path.dirname(solve_out_path), "EffectiveMassResults.xlsx")

# Initialize data structures for multiple sheets
x_direction_data = {
    "MODE": [],
    "FREQUENCY": [],
    "RATIO": []
}

y_direction_data = {
    "MODE": [],
    "FREQUENCY": [],
    "RATIO": []
}

z_direction_data = {
    "MODE": [],
    "FREQUENCY": [],
    "RATIO": []
}

rot_x_direction_data = {
    "MODE": [],
    "FREQUENCY": [],
    "RATIO": []
}

rot_y_direction_data = {
    "MODE": [],
    "FREQUENCY": [],
    "RATIO": []
}

rot_z_direction_data = {
    "MODE": [],
    "FREQUENCY": [],
    "RATIO": []
}

# Data for the "Effective Mass Summary" section
mass_summary_data = {
    "MODE": [],
    "FREQUENCY": [],
    "MODAL MASS": [],
    "KENE": [],
    "X-DIR": [],
    "RATIO% X": [],
    "Y-DIR": [],
    "RATIO% Y": [],
    "Z-DIR": [],
    "RATIO% Z": []
}

# Reading the solve.out file
with open(solve_out_path, 'r') as file:
    lines = file.readlines()

# Flags to identify sections of the file
in_x_direction = False
in_y_direction = False
in_z_direction = False
in_rotx_direction = False
in_roty_direction = False
in_rotz_direction = False
in_effective_mass_section = False

# Process each line in the file
for line in lines:
    # Check for section start keywords
    if "PARTICIPATION FACTOR CALCULATION *****  X  DIRECTION" in line:
        in_x_direction = True
        in_y_direction = in_z_direction = in_rotx_direction = in_roty_direction = in_rotz_direction = in_effective_mass_section = False
        continue
    elif "PARTICIPATION FACTOR CALCULATION *****  Y  DIRECTION" in line:
        in_y_direction = True
        in_x_direction = in_z_direction = in_rotx_direction = in_roty_direction = in_rotz_direction = in_effective_mass_section = False
        continue
    elif "PARTICIPATION FACTOR CALCULATION *****  Z  DIRECTION" in line:
        in_z_direction = True
        in_x_direction = in_y_direction = in_rotx_direction = in_roty_direction = in_rotz_direction = in_effective_mass_section = False
        continue
    elif "PARTICIPATION FACTOR CALCULATION *****ROTX DIRECTION" in line:
        in_rotx_direction = True
        in_x_direction = in_y_direction = in_z_direction = in_roty_direction = in_rotz_direction = in_effective_mass_section = False
        continue
    elif "PARTICIPATION FACTOR CALCULATION *****ROTY DIRECTION" in line:
        in_roty_direction = True
        in_x_direction = in_y_direction = in_z_direction = in_rotx_direction = in_rotz_direction = in_effective_mass_section = False
        continue
    elif "PARTICIPATION FACTOR CALCULATION *****ROTZ DIRECTION" in line:
        in_rotz_direction = True
        in_x_direction = in_y_direction = in_z_direction = in_rotx_direction = in_roty_direction = in_effective_mass_section = False
        continue
    elif "***** MODAL MASSES, KINETIC ENERGIES, AND TRANSLATIONAL EFFECTIVE MASSES SUMMARY *****" in line:
        in_effective_mass_section = True
        in_x_direction = in_y_direction = in_z_direction = in_rotx_direction = in_roty_direction = in_rotz_direction = False
        continue

    # Stop capturing data when section ends
    if "sum" in line:
        in_x_direction = in_y_direction = in_z_direction = in_rotx_direction = in_roty_direction = in_rotz_direction = in_effective_mass_section = False

    # Extract data for each direction
    columns = line.split()
    if len(columns) >= 8:
        if in_x_direction:
            # Skip the header row if it's duplicated
            if columns[0] == 'MODE':
                continue
            x_direction_data["MODE"].append(columns[0])
            x_direction_data["FREQUENCY"].append(columns[1])
            x_direction_data["RATIO"].append(columns[4])
        elif in_y_direction:
            if columns[0] == 'MODE':
                continue
            y_direction_data["MODE"].append(columns[0])
            y_direction_data["FREQUENCY"].append(columns[1])
            y_direction_data["RATIO"].append(columns[4])
        elif in_z_direction:
            if columns[0] == 'MODE':
                continue
            z_direction_data["MODE"].append(columns[0])
            z_direction_data["FREQUENCY"].append(columns[1])
            z_direction_data["RATIO"].append(columns[4])
        elif in_rotx_direction:
            if columns[0] == 'MODE':
                continue
            rot_x_direction_data["MODE"].append(columns[0])
            rot_x_direction_data["FREQUENCY"].append(columns[1])
            rot_x_direction_data["RATIO"].append(columns[4])
        elif in_roty_direction:
            if columns[0] == 'MODE':
                continue
            rot_y_direction_data["MODE"].append(columns[0])
            rot_y_direction_data["FREQUENCY"].append(columns[1])
            rot_y_direction_data["RATIO"].append(columns[4])
        elif in_rotz_direction:
            if columns[0] == 'MODE':
                continue
            rot_z_direction_data["MODE"].append(columns[0])
            rot_z_direction_data["FREQUENCY"].append(columns[1])
            rot_z_direction_data["RATIO"].append(columns[4])
        elif in_effective_mass_section and len(columns) >= 11:
            # Skip the header row if found
            if columns[0] == 'MODE':
                continue
            # Extract the data for the Effective Mass Summary
            mass_summary_data["MODE"].append(columns[0])
            mass_summary_data["FREQUENCY"].append(columns[1])
            mass_summary_data["MODAL MASS"].append(columns[2])
            mass_summary_data["KENE"].append(columns[3])
            mass_summary_data["X-DIR"].append(columns[5])
            mass_summary_data["RATIO% X"].append(columns[6])
            mass_summary_data["Y-DIR"].append(columns[7])
            mass_summary_data["RATIO% Y"].append(columns[8])
            mass_summary_data["Z-DIR"].append(columns[9])
            mass_summary_data["RATIO% Z"].append(columns[10])

# Create DataFrames for each sheet
df_x = pd.DataFrame(x_direction_data)
df_y = pd.DataFrame(y_direction_data)
df_z = pd.DataFrame(z_direction_data)
df_rotx = pd.DataFrame(rot_x_direction_data)
df_roty = pd.DataFrame(rot_y_direction_data)
df_rotz = pd.DataFrame(rot_z_direction_data)
df_mass_summary = pd.DataFrame(mass_summary_data)

# Create a summary DataFrame that consolidates the sums of RATIO from each direction
summary_data = {
    "MODE": df_x["MODE"],
    "FREQUENCY": df_x["FREQUENCY"],
    "X Direction Ratio": df_x["RATIO"],
    "Y Direction Ratio": df_y["RATIO"],
    "Z Direction Ratio": df_z["RATIO"],
    "Rot X Ratio": df_rotx["RATIO"],
    "Rot Y Ratio": df_roty["RATIO"],
    "Rot Z Ratio": df_rotz["RATIO"],
    "X-DIR": df_mass_summary["X-DIR"],
    "RATIO% X": df_mass_summary["RATIO% X"],
    "Y-DIR": df_mass_summary["Y-DIR"],
    "RATIO% Y": df_mass_summary["RATIO% Y"],
    "Z-DIR": df_mass_summary["Z-DIR"],
    "RATIO% Z": df_mass_summary["RATIO% Z"]
}

df_summary = pd.DataFrame(summary_data)

# Write data to Excel with multiple sheets
with pd.ExcelWriter(output_excel_path, engine='xlsxwriter') as writer:
    df_x.to_excel(writer, sheet_name='X Direction', index=False)
    df_y.to_excel(writer, sheet_name='Y Direction', index=False)
    df_z.to_excel(writer, sheet_name='Z Direction', index=False)
    df_rotx.to_excel(writer, sheet_name='Rot X Direction', index=False)
    df_roty.to_excel(writer, sheet_name='Rot Y Direction', index=False)
    df_rotz.to_excel(writer, sheet_name='Rot Z Direction', index=False)
    df_mass_summary.to_excel(writer, sheet_name='Effective Mass Summary', index=False)
    df_summary.to_excel(writer, sheet_name='Summary', index=False)

# Load the newly created Excel file with openpyxl to add formatting (highlighting max values)
wb = load_workbook(output_excel_path)
ws = wb["Summary"]

# Define the green fill for highlighting
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Iterate over columns C to N (those containing the ratios) to highlight max values
for col in range(3, ws.max_column + 1):
    # Find the max value in this column
    column_values = [cell for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2, values_only=True)]
    max_value = max([item for sublist in column_values for item in sublist if item is not None])  # Flattening and ignoring None values

    # Apply the green fill to cells with the max value
    for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
        if cell[0].value == max_value:
            cell[0].fill = green_fill

# Save the workbook with formatting applied
wb.save(output_excel_path)

print(f"Results extracted, summary created, and saved to {output_excel_path}")
